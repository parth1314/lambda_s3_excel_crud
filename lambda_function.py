import json
import openpyxl
import boto3
from io import BytesIO
import logging

# Initialize logger
logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client('s3')

def lambda_handler(event, context):
    logger.info(f"Received event: {json.dumps(event)}")
    
    bucket_name = 'myexlbucket' 
    file_key = 'excelfile.xlsx'  

    try:
        
        if event.get("body"):
            body = json.loads(event["body"])
        else:
            return {
                'statusCode': 400,
                'body': json.dumps({'message': 'Invalid request, no body found'})
            }
        
        
        operation = body.get('operation')
        data = body.get('data')

        logger.info(f"Operation: {operation}")
        logger.info(f"Data: {data}")

        
        logger.info("Downloading the Excel file from S3")
        s3_response = s3.get_object(Bucket=bucket_name, Key=file_key)
        file_content = s3_response['Body'].read()

        
        logger.info("Loading the Excel file")
        workbook = openpyxl.load_workbook(filename=BytesIO(file_content))
        appendix_c_sheet = workbook['Appendix C-Data']
        appendix_e_sheet = workbook['Appendix E-Data']
        assessment_findings_sheet = workbook['Assessment Findings']

        if operation == 'create':
            if data['appendix'] == 'C':
                
                last_row = appendix_c_sheet.max_row
                next_serial_number = last_row - 2  
                next_cc_reference = f'CC-{next_serial_number}'

                
                new_row = [next_serial_number, next_cc_reference] + data['values']
                appendix_c_sheet.append(new_row)
                logger.info(f"Added row to Appendix C: {new_row}")
            elif data['appendix'] == 'E':
                
                last_row = appendix_e_sheet.max_row
                next_serial_number = last_row - 2  
                next_ca_reference = f'CA-{next_serial_number}'

                
                new_row = [next_serial_number, next_ca_reference] + data['values']
                appendix_e_sheet.append(new_row)
                logger.info(f"Added row to Appendix E: {new_row}")

        elif operation == 'read':
           
            if data['appendix'] == 'C':
                sheet = appendix_c_sheet
            elif data['appendix'] == 'E':
                sheet = appendix_e_sheet

            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(row)
            return {
                'statusCode': 200,
                'body': json.dumps(rows)
            }

        elif operation == 'update':
            
            if data['appendix'] == 'C':
                sheet = appendix_c_sheet
            elif data['appendix'] == 'E':
                sheet = appendix_e_sheet

            row = data['row']
            col = data['col']
            value = data['value']
            sheet.cell(row=row, column=col).value = value
            logger.info(f"Updated cell ({row}, {col}) to {value}")

        elif operation == 'delete':
            
            if data['appendix'] == 'C':
                sheet = appendix_c_sheet
            elif data['appendix'] == 'E':
                sheet = appendix_e_sheet

            row = data['row']
            sheet.delete_rows(row)
            logger.info(f"Deleted row {row}")

        
        logger.info("Saving the updated Excel file back to S3")
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        s3.put_object(Bucket=bucket_name, Key=file_key, Body=output)
        logger.info(f"File saved back to S3: {file_key}")

        return {
            'statusCode': 200,
            'body': json.dumps({'message': 'Operation successful'})
        }

    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        return {
            'statusCode': 500,
            'body': json.dumps({'message': 'Error', 'error': str(e)})
        }
